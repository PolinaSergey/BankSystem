from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from django.contrib import admin
from django import forms
from django.http import HttpResponse

from .models import *


class CreditContractAdminForm(forms.ModelForm):
    def __init__(self, *args, **kwargs):
        super(CreditContractAdminForm, self).__init__(*args, **kwargs)
        self.fields['credit_program'].queryset = CreditProgram.objects.filter(active=True)

    class Meta:
        fields = ('credit_program','client', 'start_date', 'amount')

    def save(self, commit=True):
        credit = super(CreditContractAdminForm, self).save(commit=False)

        credit.days_left = 30 * credit.credit_program.duration
        credit.main_account = Account.objects.create(
            code="2400",
            activity="active",
            debit=0,
            credit=0,
            balance=0,
            name='Основной по кредиту {}'.format(credit.client),
            currency=credit.credit_program.currency,
            client=credit.client,
        )
        credit.percent_account = Account.objects.create(
            code="2400",
            activity="active",
            debit=0,
            credit=0,
            balance=0,
            name='Процентный по крелиту {}'.format(credit.client),
            currency=credit.credit_program.currency,
            client=credit.client,
        )

        bank_account = Account.objects.get(code='7327', currency=credit.credit_program.currency)

        with atomic():
            bank_account.withdrawal(credit.amount)
            credit.main_account.refill(credit.amount)

        cash_account = Account.objects.get(code='1010', currency=credit.credit_program.currency)

        with atomic():
            credit.main_account.withdrawal(credit.amount)
            cash_account.refill(credit.amount)

        cash_account.withdrawal(credit.amount)

        return credit


class CreditContractAdmin(admin.ModelAdmin):
    form = CreditContractAdminForm

    list_display = ('number', 'credit_program', 'client', 'start_date', 'days_left', 'end_date',
                    'amount', 'main_account', 'percent_account', 'closed')

    actions = ['close_day', 'export_to_xlsx']

    def close_day(self, request, queryset):
        for credit_contr in queryset.filter(closed=False):
            credit_contr.close_day()

    def export_to_xlsx(self, request, queryset):
        for credit_contr in queryset:

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = 'attachment; filename={date}-{number}.xlsx'.format(
                date=datetime.now().strftime('%Y-%m-%d'), number=str(credit_contr)
            )
            workbook = Workbook()
            workbook.remove(workbook.active)

            header_font = Font(name='Calibri', bold=True)
            centered_alignment = Alignment(horizontal='center')
            border_bottom = Border(
                bottom=Side(border_style='medium', color='FF000000'),
            )
            wrapped_alignment = Alignment(
                vertical='top',
                wrap_text=True
            )
            columns = [
                ('Month', 10),
                ('Amount', 30),
            ]
            worksheet = workbook.create_sheet()
            row_num = 1


            for col_num, (column_title, column_width) in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title
                cell.font = header_font
                cell.border = border_bottom
                cell.alignment = centered_alignment
                column_letter = get_column_letter(col_num)
                column_dimensions = worksheet.column_dimensions[column_letter]
                column_dimensions.width = column_width

            if credit_contr.credit_program.credit_type.name == 'Аннуитентный платеж':
                i = credit_contr.credit_program.percent / 100
                n = credit_contr.credit_program.duration
                k = (i * ((i + 1) ** n)) / (((1 + i) ** n) - 1)
                month_percent_proceeds = credit_contr.amount * Decimal(k) - (credit_contr.amount / credit_contr.credit_program.duration)
                day_proceeds = round(month_percent_proceeds / 30, 2)
                month_proceeds = day_proceeds * 30 + round(credit_contr.amount / credit_contr.credit_program.duration, 2)
                proceeds = [round(month_proceeds, 2)] * n

            else:
                proceeds = []
                for left_month in range(credit_contr.credit_program.duration, 0, -1):
                    left_amount = credit_contr.amount / credit_contr.credit_program.duration * left_month
                    day_proceeds = left_amount * Decimal(credit_contr.credit_program.percent / 100) / 30
                    day_proceeds = round(day_proceeds, 2)
                    month_proceeds = day_proceeds * 30 + \
                                     round(credit_contr.amount / credit_contr.credit_program.duration, 2)
                    proceeds.append(month_proceeds)

            for month, amount in enumerate(proceeds, 1):
                row_num += 1

                row = [
                    (month, 'Normal'),
                    (amount, 'Normal'),
                ]

                for col_num, (cell_value, cell_format) in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = cell_value
                    cell.style = cell_format
                    cell.alignment = wrapped_alignment

            worksheet.freeze_panes = worksheet['A2']

            workbook.save(response)

            return response


admin.site.register(CreditProgram)
admin.site.register(CreditContract, CreditContractAdmin)
admin.site.register(CreditType)
